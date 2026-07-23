import openpyxl
import pandas as pd
import numpy as np

wb = openpyxl.load_workbook('/mnt/user-data/uploads/_FINAL_MASTER_WITH_NEW_FORM.xlsm', data_only=True, keep_vba=True)
ws = wb['---  M A I N  ---']

headers = [ws.cell(row=8, column=c).value for c in range(1, 65)]

rows = []
for r in range(9, 278):
    rows.append([ws.cell(row=r, column=c).value for c in range(1, 65)])

df = pd.DataFrame(rows, columns=headers)

# There are duplicate-named columns (BUY DATE x2, IRR x2, COST IN x2, EXPENSE IN x2, OUT x2, HALF DATE x2)
# these are XIRR cash-flow helper columns (cols 45-51 unleveraged block, 52-58 leveraged block) - redundant
# with core columns already captured (PURCHASE DATE/SOLD DATE/PURCHASE PRICE/REHAB/SALES PROCEEDS/CASH IRR/LEVERAGED IRR)
# Drop them along with the two unnamed trailing helper date columns (59,60) and the constant col (64)
df.columns = [f"{h}__{i}" if headers.count(h) > 1 else h for i, h in enumerate(headers)]

keep_map = {
    'ADDRESS':'address','STATE':'state','PURCHASE  DATE':'purchase_date','CONTRACT PRICE':'contract_price',
    'TYPE':'seller_channel','PRODUCT TYPE':'product_type','PARTNERSHIP':'partnership','AUCTION VENDOR__7':'sourcing_vendor',
    'PURCHASE PRICE':'purchase_price','REHAB,INTEREST':'rehab_and_interest','TOTAL COST':'total_cost',
    'LOAN AMOUNT':'loan_amount','TOTAL CASH REQD':'total_cash_required','SOLD DATE':'sold_date',
    'SALES PRICE':'sales_price','SALES  CLOSING COSTS':'sales_closing_costs','SALES CLOSING %':'sales_closing_pct',
    'SALES PROCEEDS':'sales_proceeds','ACTUAL PROFIT':'actual_profit','TOTAL DAYS (DOM)':'days_on_market',
    'CASH IRR':'unleveraged_irr','LEVERAGED IRR':'leveraged_irr','BROKER CODE':'broker',
    'STRAIGHT ROI':'straight_roi','ANNUALIZED ROI':'annualized_roi','BPO VALUE':'bpo_value',
    'BPO VARIATION':'bpo_variation','PROFIT ESTIMATE':'profit_estimate','PROFIT VARIATION':'profit_variation',
    'REHAB ESTIMATE':'rehab_estimate','BROKER REHAB ACTUAL':'rehab_actual_broker','REHAB DIFFERENTIAL':'rehab_differential',
    'BUY SIDE COST ESTIMATE':'buy_side_cost_estimate','TOTAL COST ESTIMATE':'total_cost_estimate',
    'TOTAL COST ACTUAL W/REHAB':'total_cost_actual_w_rehab','DEBT SERVICE':'debt_service',
    'NOTES':'notes','Note 2':'note_2','UNDER CONTRACT':'under_contract_flag','LISTED':'listed_flag',
    'YEAR':'purchase_year','SQUARE FOOTAGE':'square_footage','AUCTION VENDOR__61':'closing_vendor',
    'BURDEN PROFIT':'burden_profit'
}
df2 = df[list(keep_map.keys())].rename(columns=keep_map)

print(df2.shape)
print(df2.dtypes)
df2.to_pickle('raw_selected.pkl')
